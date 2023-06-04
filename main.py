import os
import re
import sys
import dateparser
import codecs
import pandas as pd

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QStandardItem, QStandardItemModel, QPixmap
from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QTableView, QTableWidget,QTableWidgetItem, QFileDialog, QSplitter, QTreeView, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QFrame, QLabel, QComboBox,QPushButton,QPlainTextEdit

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


class MainWindow(QMainWindow):

    global_png_dict = {}    # 当前所有图片路径字典
    global_now_png = ""     # 当前图片路径
    current_png_list = []    # 当前图片路径集合
    current_png_index = 0 # 记录当前图片路径在当前图片路径集合中的位置

    def __init__(self):
        super().__init__()

        self.setWindowTitle("测试")
        self.initMenuBar()
        self.initMainWindows()
        self.showMaximized()
        self.root_folder_path = ""
        self.search_input.textChanged.connect(self.filterTreeView)

    def initMenuBar(self):
        """
        初始化菜单栏
        """
        # 创建菜单栏
        menu_bar = self.menuBar()

        # 创建文件菜单
        file_menu = menu_bar.addMenu("文件")
        open_folder_action = QAction("打开文件夹", self)
        open_folder_action.triggered.connect(self.openFolder)
        file_menu.addAction(open_folder_action)
        file_menu.addAction("保存")
        file_menu.addAction("退出")

        # 创建编辑菜单
        edit_menu = menu_bar.addMenu("编辑")
        edit_menu.addAction("复制")
        edit_menu.addAction("剪切")
        edit_menu.addAction("粘贴")

        # 创建帮助菜单
        help_menu = menu_bar.addMenu("帮助")
        help_menu.addAction("查看帮助")
        help_menu.addAction("关于")

    def initMainWindows(self):
        """
        初始化主窗口
        """
        splitter = QSplitter(self)

        # 创建左侧的TreeView控件
        self.tree_frame = QFrame()
        self.tree_frame.setFrameShape(QFrame.StyledPanel)  # 设置边框样式
        tree_layout = QVBoxLayout(self.tree_frame)
        self.search_input = QLineEdit()
        self.tree_view = QTreeView()
        tree_layout.addWidget(self.search_input)
        tree_layout.addWidget(self.tree_view)
        splitter.addWidget(self.tree_frame)
        self.tree_view.clicked.connect(self.treeItemClicked) # 点击事件


        # 创建右侧的TabWidget控件
        self.tab_widget = QTabWidget(self)

        # 创建”概述“页控件
        self.overview_textedit = QPlainTextEdit()
        self.tab_widget.addTab(self.overview_textedit, "概述")

        # self.tab_widget.addTab(self.table_widget, "概述")

        # 创建“截屏”页控件
        screenshot_widget = QWidget()
        screenshot_layout = QVBoxLayout(screenshot_widget)

        # 添加显示图片的控件
        self.image_label = QLabel()
        self.image_label.setFixedSize(1280, 720)
        screenshot_layout.addWidget(self.image_label)

        # 添加下拉框，上下也按钮和显示页码的控件 的水平布局
        screenshot_control_layout = QHBoxLayout()
        screenshot_layout.addLayout(screenshot_control_layout)

        # 添加下拉框
        step_switch_label = QLabel("步骤切换")
        interface_switch_label = QLabel("界面切换")
        self.step_switch_dropdown = QComboBox()
        self.interface_switch_dropdown = QComboBox()
        screenshot_control_layout.addWidget(step_switch_label)
        screenshot_control_layout.addWidget(self.step_switch_dropdown)
        screenshot_control_layout.addWidget(interface_switch_label)
        screenshot_control_layout.addWidget(self.interface_switch_dropdown)
        self.step_switch_dropdown.currentIndexChanged.connect(self.on_cbstep_changed)
        self.interface_switch_dropdown.currentIndexChanged.connect(self.on_cbinterface_changed)

        # 添加上下页按钮及显示页码的控件
        self.prev_button = QPushButton("上一页")
        self.next_button = QPushButton("下一页")
        self.page_label = QLabel()
        screenshot_control_layout.addWidget(self.prev_button)
        screenshot_control_layout.addWidget(self.page_label)
        screenshot_control_layout.addWidget(self.next_button)

        self.tab_widget.addTab(screenshot_widget, "截屏")
        self.prev_button.clicked.connect(self.previous_image)
        self.next_button.clicked.connect(self.next_image)


        self.tab_widget.addTab(QWidget(), "录波")

        # 创建“报文”页控件
        self.message_tab = QWidget()
        self.tab_widget.addTab(self.message_tab, "报文")
        # 创建报文表格控件
        layout = QVBoxLayout(self.message_tab)
        self.table = QTableWidget()
        layout.addWidget(self.table)

        splitter.addWidget(self.tab_widget)

        # 设置QSplitter的布局方式为水平布局
        splitter.setOrientation(Qt.Horizontal)

        # 设置TreeView的宽度为TabWidget宽度的1/5
        splitter.setSizes([round(self.width() / 5), round(self.width() * 4 / 5)])

        self.setCentralWidget(splitter)

        # 增加QTablew

    def openFolder(self):
        """
        打开文件夹, 用于加载目录结构
        """
        folder_dialog = QFileDialog()
        folder_dialog.setFileMode(QFileDialog.DirectoryOnly)
        folder_dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        folder_path = folder_dialog.getExistingDirectory(self, "选择文件夹", ".")
        if folder_path:
            print("选择的文件夹路径:", folder_path)
            self.root_folder_path = folder_path  # 保存文件夹路径
            self.loadTreeView(folder_path)

    def loadTreeView(self, folder_path):
        model = QStandardItemModel()

        # 加载根目录
        root_item = QStandardItem(os.path.basename(folder_path))
        root_item.setData(folder_path, Qt.UserRole)  # 设置根节点的路径属性
        root_item.setEditable(False)  # 设置根节点不可编辑
        model.appendRow(root_item)

        dir_projects = self.loadProjects(folder_path)
        self.loadSubDirectories(root_item, dir_projects)

        self.tree_view.setModel(model)
        self.tree_view.header().setVisible(False)
        self.tree_view.expandAll()

    def loadSubDirectories(self, parent_item, projects_info):
        """
        递归加载子目录
        """
        if not projects_info:
            return None
        else:
           for first_key in projects_info.keys():
               # 加载第一层级目录，工程大项
               item = QStandardItem(first_key)
               item.setData(first_key, Qt.UserRole)
               parent_item.appendRow(item)

               sort_keys = sorted(projects_info[first_key].keys(), key=self.custom_sort)

               for second_key in sort_keys:
                   # 加载第二层级目录，工程小项及路径属性
                   project_path = projects_info[first_key][second_key]
                   sub_Item = QStandardItem(second_key)
                   sub_Item.setData(project_path, Qt.UserRole)
                   item.appendRow(sub_Item)

    def filterTreeView(self, keyword):
        """
        模糊匹配目录树
        """
        if keyword:
            root_item = self.tree_view.model().item(0)
            if root_item:
                for i in range(root_item.rowCount()):
                    item = root_item.child(i)
                    self.filterItem(item, keyword)
        else:
            self.loadTreeView(self.root_folder_path)  # 加载最初的目录树

    def filterItem(self, item, keyword):
        """
        递归过滤目录树节点
        """
        if self.matchItem(item, keyword):
            index = item.index()
            self.tree_view.setRowHidden(index.row(), index.parent(), False)  # 设置节点可见
            self.tree_view.expand(index.parent())
            return True

        for i in range(item.rowCount()):
            child_item = item.child(i)
            if self.filterItem(child_item, keyword):
                index = child_item.index()
                self.tree_view.setRowHidden(index.row(), index.parent(), False)  # 设置节点可见
                self.tree_view.expand(index.parent())
                return True

        index = item.index()
        self.tree_view.setRowHidden(index.row(), index.parent(), True)  # 设置节点隐藏
        return False

    def matchItem(self, item, keyword):
        """
        判断节点是否与关键词匹配
        """
        item_text = item.text()
        if re.search(keyword, item_text, re.IGNORECASE):
            return True

        return False

    def loadProjects(self, folder_path):
        """
        加载目录下的所有工程信息
        """

        result = {}
        pattern = r'\d+\.\d+\.\d+'

        if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
            return result

        # 遍历目录下的所有文件夹 找到符合正则表达式的试验文件夹
        for root, dirs, files in os.walk(folder_path):
            for dir_name in dirs:
                if re.match(pattern, dir_name):
                    # 10.1.1 => 10 and 1.1 试验大组和试验小组拆分
                    parts = dir_name.split(".", 1)
                    first_dir = parts[0]
                    secrond_dir = parts[1]
                    secrond_result = {}

                    # 如果实验大组存在，判断实验小组是否存在，存在则比较路径，不存在则添加
                    if first_dir in result:
                        if secrond_dir in result[first_dir]:
                            path1 = result[first_dir][secrond_dir]
                            path2 = os.path.join(root, dir_name)

                            result_path = self.compareDirs(path1, path2)
                            new_path = result_path if result_path is not None else path1

                            result[first_dir][secrond_dir] = new_path.replace("\\", "/")
                        else:
                            dir_path = os.path.join(root, dir_name)
                            result[first_dir][secrond_dir] = dir_path.replace("\\", "/")
                    else:
                        dir_path = os.path.join(root, dir_name)
                        secrond_result[secrond_dir] = dir_path.replace("\\", "/")
                        result[first_dir] = secrond_result

        return result

    def compareDirs(self, path1, path2):
        """
        比较两个目录，返回最新的目录
        """

        # 截取目录路径中的年月信息
        dir1_parent = os.path.dirname(os.path.dirname(os.path.dirname(path1)))
        dir2_parent = os.path.dirname(os.path.dirname(os.path.dirname(path2)))

        dir1_year_month = os.path.basename(dir1_parent)
        dir2_year_month = os.path.basename(dir2_parent)

        year_month1 = self.parse_year_month(dir1_year_month)
        year_month2 = self.parse_year_month(dir2_year_month)

        # 截取目录路径中的试验编号信息
        dir1_number = os.path.basename(os.path.dirname(os.path.dirname(path1)))
        dir2_number = os.path.basename(os.path.dirname(os.path.dirname(path2)))

        # 比较日期和试验编号，如果日期不同，返回日期最新的目录，如果日期相同，返回试验编号最大的目录
        if year_month1 != year_month2:
            return path1 if year_month1 > year_month2 else path2
        else:
            return path1 if dir1_number > dir2_number else path2

    def parse_year_month(self, year_month):
        """
        解析年月信息
        """
        try:
            date = dateparser.parse(year_month, settings={"DATE_ORDER": "YMD"})
            return date.strftime("%Y-%m") if date else None
        except:
            return None

    def custom_sort(self, x):
        """自定义排序"""
        parts = x.split('.')
        return (int(parts[0]), int(parts[1]))

    def treeItemClicked(self, index):
        """树节点点击事件(二级节点，加载面板数据，状态栏显示当前文件路径)
        """
        item = self.tree_view.model().itemFromIndex(index)
        if item.parent() and not item.hasChildren():
            # 仅处理二级节点（试验小项）的点击事件
            file_path = item.data(Qt.UserRole)
            self.global_png_dict ={}
            self.global_last_png = []
            self.global_next_png = []
            self.global_now_png = ""
            self.global_png_dict = self.load_screenshots_path(file_path)
            self.load_excel_data(file_path)
            self.generate_step_options()
            self.load_data_from_file(file_path)

    def loadExcelData(self, excel_data):
        """
        将 Excel 数据加载到表格控件中
        @param excel_data: Excel 数据
        """
        # 清除表格中的数据
        widget = self.tab_widget.widget(0)
        for child in widget.children():
            if isinstance(child, QTableView):
                child.setParent(None)

        model = ExcelTableModel()
        table_view = QTableView()
        table_view.setModel(model)
        widget.layout().addWidget(table_view)

        for row in range(len(excel_data)):
            for col in range(len(excel_data[row])):
                item = QStandardItem(excel_data[row][col])
                model.setItem(row, col, item)

        table_view.resizeColumnsToContents()
        table_view.resizeRowsToContents()
        table_view.horizontalHeader().setStretchLastSection(True)
        table_view.verticalHeader().setSectionResizeMode(QTableView.ResizeToContents)

    def load_excel_data(self,file_path):
        # 拆解文件路径，获取文件名
        project_name = os.path.basename(file_path)
        root_path = os.path.dirname(os.path.dirname(file_path))
        excel_path = os.path.normpath(os.path.join(root_path, "试验说明"))

        if not os.path.exists(excel_path):
            return None

        project_excel_path = self.join_excel_path(excel_path, project_name)

        wb = load_workbook(project_excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        overview_text = ""

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            overview_text += f"Sheet: {sheet_name}\n\n"

            for row in sheet.iter_rows(values_only=True):
                overview_text += "\t".join(str(cell) for cell in row) + "\n"

            overview_text += "\n"

        self.overview_textedit.setPlainText(overview_text)

    def load_excel_data1(self, file_path):
        # 拆解文件路径，获取文件名
        project_name = os.path.basename(file_path)
        root_path = os.path.dirname(os.path.dirname(file_path))
        excel_path = os.path.normpath(os.path.join(root_path, "试验说明"))

        if not os.path.exists(excel_path):
            return None

        project_excel_path = self.join_excel_path(excel_path, project_name)
        workbook = load_workbook(project_excel_path, read_only=False, data_only=True)
        sheet_names = workbook.sheetnames

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]

            rows = sheet.max_row
            cols = sheet.max_column

            self.table_widget.setRowCount(rows)
            self.table_widget.setColumnCount(cols)

            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell = sheet.cell(row=row, column=col)

                    # 处理合并单元格
                    if cell.coordinate in sheet.merged_cells:
                        continue

                    # 处理图片信息
                    if isinstance(cell, Image):
                        img = cell.image
                        img_data = img._data
                        pixmap = QPixmap()
                        pixmap.loadFromData(img_data)
                        item = QTableWidgetItem()
                        item.setData(0, pixmap)
                        self.table_widget.setItem(row - 1, col - 1, item)
                    else:
                        cell_value = cell.value
                        item = QTableWidgetItem(str(cell_value))
                        self.table_widget.setItem(row - 1, col - 1, item)

            # 处理合并行和合并列
            for range_string in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = range_string.bounds
                if min_row != max_row or min_col != max_col:
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            item = QTableWidgetItem(str(cell_value))
                            self.table_widget.setItem(row - 1, col - 1, item)

            # 创建标签页并添加到QTabWidget
            tab = QWidget()

            self.tab_widget.addTab(tab, sheet_name)

            # 在标签页上放置QTableWidget
            table_widget = QTableWidget(tab)
            table_widget.setRowCount(rows)
            table_widget.setColumnCount(cols)

            # 将数据填充到QTableWidget中
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value
                    item = QTableWidgetItem(str(cell_value))
                    table_widget.setItem(row - 1, col - 1, item)

            tab_layout = QVBoxLayout()
            tab_layout.addWidget(table_widget)
            tab.setLayout(tab_layout)

    def join_excel_path(self, excel_path, project_name):
        for root, dirs, files in os.walk(excel_path):
            for file in files:
                if file.endswith('.xlsx') or file.endswith('.xls'):
                    if project_name == file.split('-')[0]:
                        project_excel_path = os.path.join(root, file)
                        return project_excel_path

        return None

    def generate_step_options(self):
        """生成下拉框的选项"""
        if not self.global_png_dict:
            return None

        # 清空步骤下拉框，防止重复添加，加载数据源并设置默认项
        self.step_switch_dropdown.clear()
        keys = self.global_png_dict.keys()
        self.step_switch_dropdown.addItems(keys)
        self.step_switch_dropdown.setCurrentIndex(0)

        key = next(iter(keys))
        values = self.global_png_dict[key]
        if values:
            category_keys = values.keys()
            if category_keys:
                self.interface_switch_dropdown.clear()
                self.interface_switch_dropdown.addItems(category_keys)
                # self.interface_switch_dropdown.setCurrentIndex(0)

                # 加载截屏文件
                category_key = next(iter(category_keys))
                self.current_png_list = values[category_key]
                self.current_png_index = 0
                self.update_page_label()

    def load_image(self):
        """加载截屏文件"""
        if not self.current_png_list:
            return

        current_png_path = self.current_png_list[self.current_png_index]
        self.load_screenshots_png(current_png_path)

    def load_screenshots_png(self, path):
        """根据路径加载截屏文件"""
        if not os.path.exists(path):
            return
        pixmap = QPixmap(path)
        if pixmap.isNull():
            return
        scaled_pixmap = pixmap.scaled(self.image_label.size(), aspectRatioMode=Qt.KeepAspectRatio)

        self.image_label.setPixmap(scaled_pixmap)

    def update_page_label(self):
        """更新页码信息"""
        page_text = f"{self.current_png_index + 1}/{len(self.current_png_list)}"
        self.page_label.setText(page_text)

    def load_screenshots_path(self, path):
        """加载所有截屏文件路径的字典"""
        step_png_dic = {}
        hmi_path = os.path.normpath(os.path.join(path, "HMI"))  # 拼接HMI路径

        # 遍历HMI文件夹下的所有文件夹-步骤信息
        for folder_name in os.listdir(hmi_path):
            png_dict = {}
            folder_path = os.path.normpath(os.path.join(hmi_path, folder_name))  # 拼接步骤信息路径

            # 判断当前路径是否为文件夹
            if not os.path.isdir(folder_path):
                continue

            for file_name in os.listdir(folder_path):
                if not file_name.endswith('.PNG'):  # 过滤非截屏文件
                    continue

                file_path = os.path.normpath(os.path.join(folder_path, file_name))  # 拼接截屏文件路径
                category = file_name.split('_')[0]  # 提取功能大类名称--画面切换数据源
                 # 将截屏文件路径添加到集合中
                if category not in png_dict:
                    file_paths = []
                    file_paths.append(file_path)
                    png_dict[category] = file_paths
                else :
                    value = png_dict[category]
                    value.append(file_path)
                    png_dict[category] = value  # 将截屏文件路径添加到集合中

            step_png_dic[folder_name] = png_dict

        return step_png_dic

    def previous_image(self):
        """上一页按钮点击事件"""
        if not self.current_png_list:
            return

        if self.current_png_index > 0:
            self.current_png_index -= 1
            self.load_image()
            self.update_page_label()

    def next_image(self):
        """下一页按钮点击事件"""
        if not self.current_png_list:
            return

        if self.current_png_index < len(self.current_png_list) - 1:
            self.current_png_index += 1
            self.load_image()
            self.update_page_label()

    def on_cbstep_changed(self):
        """步骤下拉框选中内容改变事件"""
        key = self.step_switch_dropdown.currentText()
        # 当发生改变的时候，联动界面切换下拉框
        if key in self.global_png_dict.keys():
            values = self.global_png_dict[key]
            if values:
                category_keys = values.keys()
                if category_keys:
                    self.interface_switch_dropdown.clear()
                    self.interface_switch_dropdown.addItems(category_keys)
                    current_category_key = self.interface_switch_dropdown.currentText()
                    self.current_png_list = values[current_category_key]
                    self.current_png_index = 0
                    self.update_page_label()

    def on_cbinterface_changed(self):
        """界面下拉框选中内容改变事件"""
        key = self.interface_switch_dropdown.currentText()
        if key:
            key_step = self.step_switch_dropdown.currentText()
            values = self.global_png_dict[key_step]
            now_path = values[key][0]
            self.load_screenshots_png(now_path)
            self.global_now_png = now_path
            self.current_png_list = values[key]
            self.current_png_index = 0

    def load_data_from_file(self, path):
        """读取报文文件内容"""

        # 拼接filename
        project_name = os.path.basename(path)
        filename = project_name + ".txt"
        filepath = os.path.normpath(os.path.join(path, filename))

        # 目前存储的文件编码格式为gb2312
        with codecs.open(filepath, 'r', 'gb2312') as file:
            lines = file.readlines()

        # 解析列头
        column_names = lines[0].strip().split('\t')

        # 设置表格行数和列数
        num_rows = len(lines) - 1  # 减去列头行
        num_cols = len(column_names)
        self.table.setRowCount(num_rows)
        self.table.setColumnCount(num_cols)

        # 设置列头
        self.table.setHorizontalHeaderLabels(column_names)

        # 设置列的水平对齐方式为左对齐
        header = self.table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignLeft)

        # 填充表格数据
        for row, line in enumerate(lines[1:], start=0):  # 从第二行开始
            data = line.strip().split('\t')
            if len(data) == num_cols:
                for col, item in enumerate(data):
                    table_item = QTableWidgetItem(item)
                    table_item.setTextAlignment(Qt.AlignLeft)  # 设置单元格内容居左对齐
                    self.table.setItem(row, col, table_item)

        # 调整表格大小以适应内容
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()

class ExcelTableModel(QStandardItemModel):
    """Excel model

    Args:
        QStandardItemModel (_type_): _description_
    """
    def __init__(self, parent=None):
        super().__init__(parent)

    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            item = self.itemFromIndex(index)
            if item:
                return item.text()

        return super().data(index, role)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self.horizontalHeaderItem(section).text()
            elif orientation == Qt.Vertical:
                return self.verticalHeaderItem(section).text()

        return super().headerData(section, orientation, role)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
