import os
from openpyxl import load_workbook

def read_excel(file_path):
    """
    读取试验概述数据excel文件
    # param file_path: 文件路径
    """
    # 拆解文件路径，获取文件名
    project_name = os.path.basename(file_path)
    root_path = os.path.dirname(os.path.dirname(file_path))
    excel_path = os.path.normpath(os.path.join(root_path, "试验说明"))

    if not os.path.exists(excel_path):
        return None

    project_excel_path = join_excel_path(excel_path, project_name)

    if project_excel_path:
        workbook = load_workbook(project_excel_path)
        sheet_names = workbook.sheetnames

        result = {}  # 将 result 设置字典类型

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            rows = sheet.max_row
            columns = sheet.max_column

            sheet_data = []
            for row in range(1, rows + 1):
                row_data = []
                for column in range(1, columns + 1):
                    cell = sheet.cell(row=row, column=column)

                    if cell.coordinate in sheet.merged_cells:
                        # 对于合并单元格，获取合并区域的左上角单元格的值
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                start_cell = merged_range.start_cell
                                row_data.append(sheet[start_cell.coordinate].value)
                                break
                    else:
                        row_data.append(cell.value)

                sheet_data.append(row_data)

            result[sheet_name] = sheet_data  # 使用 sheet_name 作为键

        return result
    else:
        return None

def join_excel_path(excel_path, project_name):
    for root, dirs, files in os.walk(excel_path):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                if project_name in file:
                    project_excel_path = os.path.join(root, file)
                    return project_excel_path

    return None
